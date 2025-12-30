# ============================
# 標準ライブラリ
# ============================
import os                                # 
import tempfile                          # 一時ファイル
import math                              # 数学関数
from typing import (                     # 型定義
    Tuple
)
import numpy as np                       # 数値処理
from operator import index               # インデックス操作関連
from xml.etree import ElementTree        # XML/SVG の DOM 解析
import logging                           # ログ
import asyncio                           # 非同期処理
from enum import Enum                    # 列挙型
from dataclasses import dataclass        # データクラス

# ============================
# サードパーティライブラリ
# ============================
from pydantic import (                    # Pydantic Model
    BaseModel,
    Field,
    computed_field,
    model_validator
)
from pydantic_core import(                # Pydantic の検証エラー
    PydanticCustomError, ValidationError
)
from svgpathtools import (                # SVG パス解析ツール
    parse_path,
    wsvg,
    Path,
    Line,
    CubicBezier,
    QuadraticBezier,
)
from shapely.geometry import (            # 形状処理（点・多角形）
    Point,
    Polygon
)
from shapely.ops import unary_union       # ジオメトリ結合
from fastapi import(                      # FastAPI
    FastAPI,
    Request,
    HTTPException,
    BackgroundTasks,
)
from fastapi.responses import(
    FileResponse,
    JSONResponse,
)
import openpyxl

# ロガーの初期化
logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )

logger = logging.getLogger()
logger.setLevel(logging.DEBUG)

# FastAPI の初期化
app = FastAPI(title="Sweater Chart Generator")

# 編み目記号のEnum
class Symbol(Enum):
    # 名前 = (番号, 記号)
    NONE     = (0, "#")

    KNIT     = (1, "")
    PURL     = (-1, "-")
    M1       = (2, "△")

    K2TOG    = (11, "＼")
    P2TOG    = (-11, "＼")

    SSK      = (21, "／")
    SSP      = (-21, "／")

    CO       = (50, "▲")
    BO       = (60, "●")
    HOLD     = (70, "■")

    _MARKER_1 = (-101, "")
    _MARKER_2 = (-102, "")

    def __init__(self, number, char):
        self.number = number
        self.char = char

    @classmethod
    def from_number(cls, num):
        for item in cls:
            if item.number == num:
                return item
        return cls.NONE

# セーターの形状のEnum
class SweaterType(str, Enum):
    CREW_NECK_SWEATER = "crew-neck-sweater"
    V_NECK_SWEATER = "v-neck-sweater"
    BOX_SWEATER = "box-sweater"
    TURTLENECK_SWEATER = "turtleneck"

    CREW_NECK_CARDIGAN = "crew-neck-cardigan"
    V_NECK_CARDIGAN = "v-neck-cardigan"
    BOX_CARDIGAN = "box-cardigan"

    CREW_NECK_VEST = "crew-neck-vest"
    V_NECK_VEST = "v-neck-vest"
    BOX_VEST = "box-vest"

    PO_CREW_NECK_VEST = "po-crew-neck-vest"
    PO_V_NECK_VEST = "po-v-neck-vest"
    PO_BOX_VEST = "po-box-vest"

class Metric(str, Enum):
    INCH = "inch"
    CM = "cm"

class Gauge(BaseModel):
    """ ゲージの幅と高さを表現するクラス """
    metric: Metric = Field(default=Metric.CM, description="寸法の単位")
    vertical: float = Field(..., description="ゲージの段数", gt=0)
    horizontal: float = Field(..., description="ゲージの目数", gt=0)
    @computed_field
    @property
    def stitch_width(self):
        if self.metric == Metric.INCH:
            return 10.16 / self.horizontal
        elif self.metric == Metric.CM:
            return 10 / self.horizontal
        return 0
    
    @computed_field
    @property
    def stitch_length(self):
        if self.metric == Metric.INCH:
            return 10.16 / self.vertical
        elif self.metric == Metric.CM:
            return 10 / self.vertical
        return 0
 

# POSTで受け取るデータを表現するクラス Pydantic Model
class SweaterDimensions(BaseModel):
    """
    長袖のセーターの寸法とゲージを定義するデータモデル
    """
    gauge: Gauge = Field(..., description="ゲージ")

    length_of_body: float = Field(..., description="着丈", gt=0)
    length_of_shoulder_drop: float = Field(..., description="肩下がり", gt=0)
    length_of_ribbed_hem: float = Field(..., description="裾ゴム編み", gt=0)

    length_of_front_neck_drop: float = Field(..., description="前襟ぐり下がり", gt=0)
    length_of_back_neck_drop: float = Field(..., description="後ろ襟ぐり下がり", gt=0)

    width_of_body: float = Field(..., description="身幅", gt=0)
    width_of_neck: float = Field(..., description="襟ぐり幅", gt=0)

    length_of_sleeve: float = Field(..., description="袖丈", gt=0)
    length_of_ribbed_cuff: float = Field(..., description="袖口ゴム編み", gt=0)

    width_of_sleeve: float = Field(..., description="袖幅", gt=0)
    width_of_cuff: float = Field(..., description="袖口幅", gt=0)

    type: SweaterType = Field(..., description="セーターの形状")

    is_odd: bool = Field(default=False, description="水平方向の目数を奇数にする")

    def __str__(self) -> str:
        lines = [
            f"SweaterDimensions:",
            f"gauge_height: {self.gauge.vertical}",
            f"gauge_width: {self.gauge.horizontal}",
            "",
            f"length_of_body: {self.length_of_body}",
            f"length_of_shoulder_drop: {self.length_of_shoulder_drop}",
            f"length_of_ribbed_hem: {self.length_of_ribbed_hem}",
            f"length_of_front_neck_drop: {self.length_of_front_neck_drop}",
            f"length_of_back_neck_drop: {self.length_of_back_neck_drop}",
            "",
            f"width_of_body: {self.width_of_body}",
            f"width_of_neck: {self.width_of_neck}",
            "",
            f"length_of_sleeve: {self.length_of_sleeve}",
            f"length_of_ribbed_cuff: {self.length_of_ribbed_cuff}",
            "",
            f"width_of_sleeve: {self.width_of_sleeve}",
            f"width_of_cuff: {self.width_of_cuff}",
            "",
            f"type: {self.type}",
            "",
            f"is_odd: {self.is_odd}",
            "",
            f"stitch_width: {self.gauge.stitch_width}",
            f"stitch_length: {self.gauge.stitch_length}",
            "",
            f"length_of_body_side: {self.length_of_body_side}",
            f"length_of_vertical_armhole: {self.length_of_vertical_armhole}",
            "",
            f"width_of_horizontal_armhole: {self.width_of_horizontal_armhole}",
            f"width_of_shoulder: {self.width_of_shoulder}",
            "",
            f"length_of_sleeve_cap: {self.length_of_sleeve_cap}",
            f"length_of_sleeve_side: {self.length_of_sleeve_side}"
        ]
        return "\n".join(lines)

    @computed_field
    @property
    def length_of_body_side(self) -> float:
        """ 脇下から裾のゴム編みの上端までの長さ"""
        return self.length_of_body - self.length_of_shoulder_drop - self.length_of_vertical_armhole - self.length_of_ribbed_hem

    @computed_field
    @property
    def length_of_vertical_armhole(self) -> float:
        """ 袖ぐりの垂直方向の長さ """
        return int(self.width_of_sleeve / self.gauge.stitch_length) * self.gauge.stitch_length

    @computed_field
    @property
    def width_of_horizontal_armhole(self) -> float:
        """ 袖ぐりの水平方向の長さ 身幅の1/10"""
        return int((self.width_of_body * 0.1) / self.gauge.stitch_width) * self.gauge.stitch_width

    @computed_field
    @property
    def width_of_shoulder(self) -> float:
        """ 肩幅(mm) 身幅から袖ぐりの水平方向と襟ぐり幅を引いた長さの1/2"""
        return (self.width_of_body - self.width_of_horizontal_armhole * 2 - self.width_of_neck) / 2

    @computed_field
    @property
    def length_of_sleeve_cap(self) -> float:
        """ 袖山の高さ(mm) 袖ぐりの水平方向の長さの2倍"""
        return int((self.width_of_horizontal_armhole * 2) / self.gauge.stitch_length) * self.gauge.stitch_length

    @computed_field
    @property
    def length_of_sleeve_side(self) -> float:
        """ 袖下(mm) 袖丈から袖山の高さ・袖口のゴム編みを引いた長さ"""
        return self.length_of_sleeve - self.length_of_sleeve_cap - self.length_of_ribbed_cuff
    
    @computed_field
    @property
    def rows_of_body(self) -> int:
        """ 着丈の段数"""
        return int(self.length_of_body / self.gauge.stitch_length)
    
    @computed_field
    @property
    def rows_of_shoulder_drop(self) -> int:
        """ 肩下がりの段数"""
        return int(self.length_of_shoulder_drop / self.gauge.stitch_length)
    
    @computed_field
    @property
    def rows_of_ribbed_hem(self) -> int:
        """ 裾のゴム編みの段数"""
        return int(self.length_of_ribbed_hem / self.gauge.stitch_length)
    
    @computed_field
    @property
    def rows_of_front_neck_drop(self) -> int:
        """ 前襟ぐり下がりの段数"""
        return int(self.length_of_front_neck_drop / self.gauge.stitch_length)
    
    @computed_field
    @property
    def rows_of_back_neck_drop(self) -> int:
        """ 後襟ぐり下がりの段数"""
        return int(self.length_of_back_neck_drop / self.gauge.stitch_length)
    
    @computed_field
    @property
    def cols_of_body(self) -> int:
        """ 身幅の目数"""
        return int(self.width_of_body / self.gauge.stitch_width)
    
    @computed_field
    @property
    def cols_of_neck(self) -> int:
        """ 襟ぐり幅の目数"""
        return int(self.width_of_neck / self.gauge.stitch_width)
    
    @computed_field
    @property
    def rows_of_sleeve(self) -> int:
        """ 袖丈の段数"""
        return int(self.length_of_sleeve / self.gauge.stitch_length)
    
    @computed_field
    @property
    def rows_of_ribbed_cuff(self) -> int:
        """ 袖口のゴム編みの段数"""
        return int(self.length_of_ribbed_cuff / self.gauge.stitch_length)
    
    @computed_field
    @property
    def cols_of_sleeve(self) -> int:
        """ 袖幅の目数"""
        return int(self.width_of_sleeve / self.gauge.stitch_width)
    
    @computed_field
    @property
    def cols_of_cuff(self) -> int:
        """ 袖口幅の目数"""
        return int(self.width_of_cuff / self.gauge.stitch_width)
    
    @computed_field
    @property
    def rows_of_body_side(self) -> int:
        """ 脇下から裾のゴム編みの上端までの段数"""
        return int(self.length_of_body_side / self.gauge.stitch_length)
    
    @computed_field
    @property
    def rows_of_vertical_armhole(self) -> int:
        """ 袖ぐりの垂直方向の段数"""
        return int(self.length_of_vertical_armhole / self.gauge.stitch_length)
    
    @computed_field
    @property
    def cols_of_horizontal_armhole(self) -> int:
        """ 袖ぐりの水平方向の目数"""
        return int(self.width_of_horizontal_armhole / self.gauge.stitch_width)
    
    @computed_field
    @property
    def cols_of_shoulder(self) -> int:
        """ 肩幅の目数"""
        return int(self.width_of_shoulder / self.gauge.stitch_width)
    
    @computed_field
    @property
    def rows_of_sleeve_side(self) -> int:
        """ 袖下から袖山の高さまでの段数"""
        return int(self.length_of_sleeve_side / self.gauge.stitch_length)
    
    @computed_field
    @property
    def rows_of_sleeve_cap(self) -> int:
        """ 袖山の段数"""
        return int(self.length_of_sleeve_cap / self.gauge.stitch_length)

    def _round_to_multiple_stitch_length(self, value) -> float:
        """ 寸法を stitch_length の整数倍に丸める"""
        return int(value / self.gauge.stitch_length) * self.gauge.stitch_length

    def _round_to_multiple_stitch_width(self, value) -> float:
        """ 寸法を stitch_width の整数倍に丸める"""
        return int(value / self.gauge.stitch_width) * self.gauge.stitch_width

    def _round_to_multiple_odd_or_even_stitch_width(self, value) -> float:
        """ 寸法を stitch_width の奇数倍または偶数倍に丸める"""
        if self.is_odd:
            return int(value / self.gauge.stitch_width / 2) * 2 * self.gauge.stitch_width + self.gauge.stitch_width
        else:
            return int(value / self.gauge.stitch_width / 2) * 2 * self.gauge.stitch_width

    def _round_to_multiple_odd_or_even_stitch_width_half(self, value) -> float:
        """ 寸法を stitch_width の(整数+1/2)倍または整数倍に丸める"""
        if self.is_odd:
            return self._round_to_multiple_stitch_width(value) + self.gauge.stitch_width / 2
        else:
            return self._round_to_multiple_stitch_width(value)

    @model_validator(mode='after')
    def _adjust_and_check_dimensions(self) -> 'SweaterDimensions':
        """寸法の調整と検証を行う"""
        
        # 編目の縦の長さの整数倍に丸めた着丈
        self.length_of_body = self._round_to_multiple_stitch_length(self.length_of_body)

        # 編目の縦の長さの整数倍に丸めた肩下がり
        self.length_of_shoulder_drop = self._round_to_multiple_stitch_length(self.length_of_shoulder_drop)

        # 編目の縦の長さの整数倍に丸めた裾のゴム編み
        self.length_of_ribbed_hem = self._round_to_multiple_stitch_length(self.length_of_ribbed_hem)

        # 編目の縦の長さの整数倍に丸めた前襟ぐり下がり
        self.length_of_front_neck_drop = self._round_to_multiple_stitch_length(self.length_of_front_neck_drop)

        # 編目の縦の長さの整数倍に丸めた後襟ぐり下がり
        self.length_of_back_neck_drop = self._round_to_multiple_stitch_length(self.length_of_back_neck_drop)

        # 編目の横の長さの奇数倍または偶数倍に丸めた身幅
        self.width_of_body = self._round_to_multiple_odd_or_even_stitch_width(self.width_of_body)

        # 編目の横の長さの半分の奇数倍または偶数倍に丸めた襟幅
        self.width_of_neck = self._round_to_multiple_odd_or_even_stitch_width(self.width_of_neck)

        # 編目の縦の長さの整数倍に丸めた袖丈
        self.length_of_sleeve = self._round_to_multiple_stitch_length(self.length_of_sleeve)

        # 編目の縦の長さの整数倍に丸めた袖口のゴム編み
        self.length_of_ribbed_cuff = self._round_to_multiple_stitch_length(self.length_of_ribbed_cuff)

        # 編目の横の長さの半分の奇数倍または偶数倍に丸めた袖幅
        self.width_of_sleeve = self._round_to_multiple_odd_or_even_stitch_width_half(self.width_of_sleeve)

        # 編目の横の長さの半分の奇数倍または偶数倍に丸めた袖口幅
        self.width_of_cuff = self._round_to_multiple_odd_or_even_stitch_width_half(self.width_of_cuff)

        # TODO 寸法値の検証
        if False:
            raise PydanticCustomError(
                'value_error', # エラータイプ
                'エラーの理由',
                {'min_length': 0} # JSONに出力したい追加情報
            )

        logger.info(f"SweaterDimensions is initialized.")
        logger.debug(self)
        return self

# 検証エラーを捕捉するための例外ハンドラー
@app.exception_handler(ValidationError)
async def validation_exception_handler(request: Request, exc: ValidationError):
    """
    検証エラーが発生した場合に実行されるハンドラー
    """
    error_details = exc.errors()
    # 失敗ログの出力
    logger.error(f"SweaterDimensions の検証に失敗しました: {error_details}")
    
    # クライアントには、詳細なエラー情報を返す
    return JSONResponse(
        status_code=422,
        content={"detail": "input value is invalid.", "errors": error_details},
    )

# シェイプ（型紙）
class Shape:
    def __init__(self, path: Path, gauge: Gauge):
        self.path = path
        self.gauge = gauge

    def __getattr__(self, name):
        # クラスにないものは Path に投げる
        return getattr(self.path, name)

    @classmethod
    def front_body_from(cls, data: SweaterDimensions) -> 'Shape':
        return cls._body_from(data, is_front=True)
    
    @classmethod
    def back_body_from(cls, data: SweaterDimensions) -> 'Shape':
        return cls._body_from(data, is_front=False)

    @classmethod
    def _body_from(cls, data, is_front:bool) -> 'Shape':
        """
        身頃の Shape を生成する

        Args:
            data (SweaterDimensions): 寸法データクラス
            is_front (bool): 前身頃または後身頃

        Returns:
            Shape: 生成された身頃のシェイプ


        """

        # 襟ぐり下がり
        length_of_neck_drop = data.length_of_back_neck_drop
        if is_front:
            length_of_neck_drop = data.length_of_front_neck_drop
            

        # 肩の水平な直線の長さ
        stitch_num_of_shoulder_drop = int(data.length_of_shoulder_drop / data.gauge.stitch_length)
        stitch_num_of_shoulder_width = int(data.width_of_shoulder / data.gauge.stitch_width)
        stitch_num_of_horizonal_shoulder_line = int((stitch_num_of_shoulder_width / stitch_num_of_shoulder_drop)/2)
        horizonal_shoulder_line_width = stitch_num_of_horizonal_shoulder_line * data.gauge.stitch_width

        # 身頃のパス
        path = parse_path(
            # 起点に移動 右の脇下
            f"M {0} {data.length_of_shoulder_drop + data.length_of_vertical_armhole} "

            # 袖ぐりの半分までの曲線
            f"c {data.width_of_horizontal_armhole},0 {data.width_of_horizontal_armhole},-{data.length_of_vertical_armhole / 2} {data.width_of_horizontal_armhole},-{data.length_of_vertical_armhole / 2} "

            # 左肩の左端までの垂直な直線
            f"v -{data.length_of_vertical_armhole / 2} "

            # 左肩の右端の少し前までの直線
            f"l {data.width_of_shoulder - horizonal_shoulder_line_width},-{data.length_of_shoulder_drop} "

            # 襟ぐりの左の上端までの水平な直線
            f"h {horizonal_shoulder_line_width} "

            # 襟ぐりの下端までの曲線
            f"c 0,{length_of_neck_drop} {data.width_of_neck / 2},{length_of_neck_drop} {data.width_of_neck / 2},{length_of_neck_drop} "

            # 襟ぐりの右の上端までの曲線
            f"c {data.width_of_neck / 2},0 {data.width_of_neck / 2},-{length_of_neck_drop} {data.width_of_neck / 2},-{length_of_neck_drop} "

            # 襟ぐりの右の上端から少し右への水平な直線
            f"h {horizonal_shoulder_line_width} "

            # 右肩の右端までの直線
            f"l {data.width_of_shoulder - horizonal_shoulder_line_width},{data.length_of_shoulder_drop} "

            # 右の襟ぐりの半分までの直線
            f"v {data.length_of_vertical_armhole / 2} "

            # 脇下までの曲線
            f"c 0,{data.length_of_vertical_armhole / 2} {data.width_of_horizontal_armhole},{data.length_of_vertical_armhole / 2} {data.width_of_horizontal_armhole},{data.length_of_vertical_armhole / 2} "

            # 裾の右の下端までの垂直な直線
            f"v {data.length_of_body_side + data.length_of_ribbed_hem} "

            # 裾の端から端までの水平な直線
            f"h {-data.width_of_body}"

            # 始点まで
            f"Z"
        )
        
        return cls(path, data.gauge)

    @classmethod
    def sleeve_from(cls, data: SweaterDimensions) -> 'Shape':
        """
        袖のシェイプ（SVG）を生成する

        Args:
            data (SweaterDimensions): 検証済みの寸法データクラス

        Returns:
            Shape: 生成された袖のシェイプ

        """

        # 袖のパス
        path = parse_path(
            # 始点 袖山の左端まで移動する
            f"M {0},{data.length_of_sleeve_cap} "

            # 袖山のトップまでの曲線
            f" c {data.width_of_sleeve/2},{0} {data.width_of_sleeve/2},{-data.length_of_sleeve_cap} {data.width_of_sleeve},{-data.length_of_sleeve_cap} "

            # 袖山の右端までの曲線
            f" c {data.width_of_sleeve/2},{0} {data.width_of_sleeve/2},{data.length_of_sleeve_cap} {data.width_of_sleeve},{data.length_of_sleeve_cap} "

            # 右の袖の上端までの斜めの直線
            f"l {data.width_of_cuff - data.width_of_sleeve},{data.length_of_sleeve_side} "

            # 右の袖の上端から下端までの垂直な直線
            f"v {data.length_of_ribbed_cuff} "

            # 袖の下端の端から端までの水平な直線
            f"h {-data.width_of_cuff * 2} "

            # 左の袖の下端のから上端での垂直な直線
            f"v {-data.length_of_ribbed_cuff} "

            # 始点まで
            f"Z"
        )

        return cls(path, data.gauge)
    
    def write_svg(self, filename: str, **kwargs):
        wsvg(self.path, filename=filename, **kwargs)


# チャート（編み図）
class Chart:
    def __init__(self, array: np.ndarray, gauge: Gauge):
        self.array = array
        self.gauge = gauge

    def __getattr__(self, name):
        # クラスにないものはnp.ndarrayに投げる
        return getattr(self.array, name)
    
    @classmethod
    def from_shape(cls, shape: Shape) -> 'Chart':
        """
        Shape の Path オブジェクトを元に Chart を生成します

        まず、Path を Shapely の Polygon オブジェクトに変換します

        次に、縦に並んだ2つのグリッド（判定範囲）の中心（判定点）がPolygonオブジェクトの
        内部に存在するかどうかを判定し二次元配列を生成します。
        左右の端で折り返し位置をずらすために、ビューボックスの中心を対称軸として
        右半分と左半分で判定範囲をずらします。

        編み目記号を挿入します。

        Args:
            shape: Shape 

        Returns:
            np.ndarray: チャートの二次元配列
        """

        logger.debug(f"<<< Generating chart from shape >>>")

        # バウンドボックスのサイズを取得する
        start_x, end_x, start_y,  end_y = shape.path.bbox()
        width = end_x - start_x
        height = end_y - start_y

        # グリッドの縦横の数を計算する
        num_grid_width = int(width / shape.gauge.stitch_width)
        num_grid_height = int(height / (shape.gauge.stitch_length))
        num_grid_width_half = int(num_grid_width / 2) #対称軸の位置

        logger.debug(
            f"array size:\n"
            f"num_grid_width={num_grid_width}\n"
            f"num_grid_height={num_grid_height}\n"
            f"num_grid_width_half={num_grid_width_half}"
        )

        # グリッドと同じ行列数の配列を生成
        array = np.zeros((num_grid_height, num_grid_width), dtype=np.int8)
        logger.debug(f"grid_array is created")

        # パス要素からポリゴンを生成
        polygon = None
        polygon_points = [] # パスを線分に分割して点を取得
        num_samples = 100 # サンプリング数
        for segment in shape.path:
            if isinstance(segment, (Line, CubicBezier, QuadraticBezier)):
                for i in range(num_samples + 1):
                    t = i / num_samples
                    p = segment.point(t)
                    polygon_points.append((p.real, p.imag))
            else: pass
        # 閉じたパスの場合のみポリゴンとして追加
        if shape.path.isclosed() and len(polygon_points) >= 3:
            try:
                # TopologyException を回避するための裏技
                polygon = Polygon(polygon_points).buffer(0)
            except Exception as e:
                logger.warning(f"Could not create polygon from path due to {e}")
        else: pass

        if not polygon:
            # 形状が一つも抽出されなかった場合
            logger.warning(f"No polygon could be created from the shape path. Returning empty chart.")
            return Chart(array, shape.gauge)
        
        # 1目の縦横の長さに対応したグリッドの中心が内側かどうかは判定する
        for y_index, y_coordinate in enumerate(np.arange(0, height, shape.gauge.stitch_length)):
            for x_index, x_coordinate in enumerate(np.arange(0, width, shape.gauge.stitch_width)):
                # 判定点
                point = Point(float(x_coordinate + shape.gauge.stitch_width / 2), float(y_coordinate + shape.gauge.stitch_length / 2))
                # 右側の判定点が結合された形状の内部にあるか判定
                if polygon.contains(point):
                    # 内部にある場合、グリッドを描画
                    array[y_index, x_index] = Symbol.KNIT.number

        result = cls(array, shape.gauge)
        result._insert_symbol()

        logger.debug(f"grid_array is generated: shape{array.shape[0]} length_of_x={array.shape[1]}")
        return result

    def _insert_symbol(self) -> np.ndarray:
        """
        1.伏止め・減らし目が適切な位置になるように、段の位置を
            a. 右上がりの段は偶数行目にあらわれる
            b. 左上がりの段は奇数行目にあらわれる
        に再配置する
        2.編み目記号を挿入する

        Returns:
            np.ndarray: 変換後の配列
        """

        # === 1. 段の位置を調整する ===

        # 1.1 開始行とステップ行数の設定
        # 右上がりの段のマーカーは奇数行、左上がりの段のマーカーは偶数行に挿入する
        odd_start_row, even_start_row = 0, 1 # 配列の行数が奇数の場合
        if self.array.shape[0] % 2 == 0:
            odd_start_row, even_start_row = 1, 0 # 配列の行数が偶数の場合

        step_rows = 2

        # 1.2 右上がりの段の上のグリッドにマーカーを挿入
        self._replace_in(
            target_array = np.array([
                [Symbol.NONE.number, Symbol.NONE.number],
                [Symbol.NONE.number, Symbol.KNIT.number]
                ]),
            replacement=Symbol._MARKER_1.number,
            replacement_position=(0,1),
            start_row=odd_start_row,
            step_rows=step_rows
        )

        # 1.3 左上がりの段の上のグリッドにマーカーを挿入
        self._replace_in(
            target_array = np.array([
                [Symbol.NONE.number, Symbol.NONE.number],
                [Symbol.KNIT.number, Symbol.NONE.number]
                ]),
            replacement=Symbol._MARKER_2.number,
            replacement_position=(0,0),
            start_row=even_start_row,
            step_rows=step_rows
        )

        # 1.4 マーカーを水平方向のSymbol.NONEに伝播させ、Symbol.NONEがなくなるまで繰り返し、最後にマーカーをSymbol.KNITに置換
        while True:
            # 変化前の配列
            prev_array = self.array.copy()

            # Symbol._MARKER_1の右にあるSymbol.NONEを見つけて置換
            self._replace_in(
                target_array=np.array([
                    [Symbol._MARKER_1.number, Symbol.NONE.number]
                    ]),
                replacement=Symbol._MARKER_1.number,
                replacement_position=(0,1)
                )

            # Symbol._MARKER_2の左にあるSymbol.NONEを見つけて置換
            self._replace_in(
                target_array=np.array([
                    [Symbol.NONE.number, Symbol._MARKER_2.number]
                    ]),
                replacement=Symbol._MARKER_2.number,
                replacement_position=(0,0)
                )
            
            # 変化がなければ終了
            if np.array_equal(prev_array, self.array):
                # マーカーを Symbol.KNIT.number に置換
                self.array[self.array == Symbol._MARKER_1.number] = Symbol.KNIT.number
                self.array[self.array == Symbol._MARKER_2.number] = Symbol.KNIT.number
                break

        # ===== 2.編み目記号を挿入する =====

        # 2.1 最上行にSymbol.NONEの行を追加
        self._insert_row_to_top(fill=Symbol.NONE.number)

        # 2.2 伏止記号の挿入
        self._replace_in(
            target_array = np.array([
                [Symbol.NONE.number],
                [Symbol.KNIT.number]
                ]),
            replacement=Symbol.BO.number,
            replacement_position=(0,0)
        )

        # 2.3 右上2目1度記号の挿入
        self._replace_in(
            target_array = np.array([
                [Symbol.KNIT.number, Symbol.BO.number]
                ]),
            replacement=Symbol.K2TOG.number,
            replacement_position=(0,1)
        )

        # 2.4 左上2目1度記号の挿入
        self._replace_in(
            target_array = np.array([
                [Symbol.BO.number, Symbol.KNIT.number]
                ]),
            replacement=Symbol.SSK.number,
            replacement_position=(0,0)
        )

        # 2.5 増目記号の挿入
        self._replace_in(
            target_array = np.array([
                [Symbol.KNIT.number, Symbol.KNIT.number],
                [Symbol.KNIT.number, Symbol.NONE.number]
                ]),
            replacement=Symbol.M1.number,
            replacement_position=(0,0)
        )
        self._replace_in(
            target_array = np.array([
                [Symbol.KNIT.number, Symbol.KNIT.number],
                [Symbol.NONE.number, Symbol.KNIT.number]
                ]),
            replacement=Symbol.M1.number,
            replacement_position=(0,1)
        )

        return self.array

    def _replace_in(self, 
                target_array: np.ndarray, 
                replacement: int, 
                replacement_position: Tuple[int, int],
                start_row: int = 0,
                step_rows: int = 1) -> np.ndarray:
        """
        指定した行範囲・ステップで target_array に一致するパターンを検索し、
        特定の1点 (replacement_position) を置換する。
        """
        rows, cols = target_array.shape
        r_y, r_x = replacement_position
        
        arr = self.array
        h, w = arr.shape
        
        # --- 1. 判定範囲の決定 ---
        # 判定を開始できる最終行のインデックス
        last_possible_row = h - rows
        
        # start_row から last_possible_row までの範囲で step_rows ごとにインデックスを作成
        if start_row > last_possible_row:
            return arr # 判定不能な場合はそのまま返す
            
        row_indices = np.arange(start_row, last_possible_row + 1, step_rows)
        
        # --- 2. マスクの作成 ---
        # マスクの形状は (選択された行数, w - cols + 1)
        mask_h = len(row_indices)
        mask_w = w - cols + 1
        final_mask = np.ones((mask_h, mask_w), dtype=bool)
        
        for y in range(rows):
            for x in range(cols):
                # 行方向は row_indices を起点に y ずらした位置を参照
                # 列方向は 0 から (w - cols + x + 1) まで
                # np.ix_ を使うと不連続なインデックス（step）での抽出がスムーズです
                selected_rows = row_indices + y
                slice_part = arr[selected_rows[:, None], np.arange(x, w - cols + x + 1)]
                
                final_mask &= (slice_part == target_array[y, x])
        
        # --- 3. 置換の実行 ---
        result = arr.copy()
        
        # 置換対象の y 座標を計算 (各判定開始行 + replacement_positionのy)
        replace_y_indices = row_indices + r_y
        # 置換対象の x 座標の範囲
        replace_x_start = r_x
        replace_x_end = w - cols + r_x + 1
        
        # mask が True の場所だけ replacement を代入
        rows_to_change, cols_to_change = np.where(final_mask)
        
        actual_y = replace_y_indices[rows_to_change]
        actual_x = replace_x_start + cols_to_change
        
        result[actual_y, actual_x] = replacement
        
        self.array = result
        return result
    
    def _insert_row_to_top(self, fill: int) -> np.ndarray:
        """
        配列の先頭に fill の行を挿入する
        """
        result = self.array.copy()
        result = np.insert(result, 0, fill, axis=0)
        self.array = result
        return result
    
    def insert_pattern_repeatedly(
        self, 
        pattern: np.ndarray, 
        start_row: int = 0, 
        end_row: int = None, # type: ignore
        start_col: int = 0, 
        end_col: int = None # type: ignore
    ):
        """
        指定された矩形範囲に、patternを繰り返し挿入する
        """
        # 範囲の終点が未指定の場合は、グリッドの端までとする
        if end_row is None:
            end_row = self.array.shape[0]
        if end_col is None:
            end_col = self.array.shape[1]

        # 挿入する領域のサイズを計算
        target_h = end_row - start_row
        target_w = end_col - start_col

        # 負の範囲やゼロサイズの場合は何もしない
        if target_h <= 0 or target_w <= 0:
            return

        # 1. 繰り返し回数を計算 (パターンのサイズで領域を割る)
        reps_y = int(np.ceil(target_h / pattern.shape[0]))
        reps_x = int(np.ceil(target_w / pattern.shape[1]))

        # 2. タイル状に並べて、ターゲットのサイズに正確に切り抜く
        # np.tile は (縦の反復数, 横の反復数) を受け取る
        tiled = np.tile(pattern, (reps_y, reps_x))
        final_patch = tiled[:target_h, :target_w]

        # 3. 指定された矩形範囲を上書き
        result = self.array.copy()
        result[start_row:end_row, start_col:end_col] = final_patch

        self.array = result
        return result
        
    def symmetrize_rows(self, start_row: int = 0, end_row: int = None, based_on_right: bool = False) -> np.ndarray: # type: ignore
        """
        チャートの指定された範囲の行を片側を基準にして左右対称にする
        """
        if end_row is None:
            end_row = self.array.shape[0]

        print(self.array.shape)

        result = self.array.copy()

        # 列数が偶数か奇数かで中心列の位置が変わる
        if self.array.shape[1] % 2 == 0:
            center_col_end = self.array.shape[1] // 2
            center_col_start = center_col_end
        else:
            center_col_end = self.array.shape[1] // 2
            center_col_start = center_col_end + 1

        if based_on_right:
            # 右側を基準に左側を反転
            result[start_row:end_row, center_col_start:] = result[start_row:end_row, :center_col_end][:, ::-1]
        else:
            # 左側を基準に右側を反転
            result[start_row:end_row, :center_col_end] = result[start_row:end_row, center_col_start:][:, ::-1]

        self.array = result
        return result

    def write_csv(self, filename: str):
        """ チャートをCSVファイルに書き出す """
        np.savetxt(
            filename,
            self.array,
            fmt='%d',
            delimiter=','
        )


class XLSX():
    def __init__(self, xlsx: openpyxl.Workbook):
        self._xlsx = xlsx

    def __getattr__(self, name):
        # クラスにないものはopenpyxl.Workbookに投げる
        return getattr(self._xlsx, name)

    @classmethod
    def from_charts(cls, charts: dict[str, Chart]) -> 'XLSX':
        # TODO

        keys = np.array(list(charts.keys()))
        num_charts = keys.shape[0]

        # 先頭のチャートからゲージを取得
        gauge = charts[keys[0]].gauge

        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise ValueError
        ws.title = "info"

        for i in range(num_charts):
            chart = charts[keys[i]]
            ws = wb.create_sheet(keys[i])
            for row in range(chart.array.shape[0]):
                for col in range(chart.array.shape[1]):
                    ws.cell(row=row+1, column=col+1).value = chart.array[row, col]

        if wb is None:
            raise ValueError
        else:
            return cls(wb)
        

    
    
@app.post("/generate_sweater_chart", response_description="generated file")
async def main(sweaterDimensions: SweaterDimensions, is_debug=False):
    """
    Pydanticモデルで受け取ったデータから生成したファイルを送信する
    
    Args:
        data (SweaterDimensions): 検証済みの寸法データクラス
    """

    file_path = generate_file(sweaterDimensions)

    background = BackgroundTasks()
    background.add_task(cleanup_file(file_path))

    try:
        # 1. コアロジックを実行し、一時ファイルのパスを取得
        file_path = generate_file(sweaterDimensions)
        
        # 2. FileResponseでファイルをクライアントに送信
        return FileResponse(
            path=file_path, 
            filename="sweater_pattern_data.csv",
            media_type="text/csv", # TODO .xlsx .pdf
            background=background # 送信後にファイルを削除するタスクを設定
        )

    except Exception as e:
        if file_path and os.path.exists(file_path):
            os.remove(file_path)
        raise HTTPException(status_code=500, detail=f"ファイル生成中にエラーが発生しました: {str(e)}")

def cleanup_file(file_path: str):
    import os
    def _cleanup():
        # ファイルが存在することを確認してから削除
        if os.path.exists(file_path):
            os.remove(file_path)
    return _cleanup
    
    if data.body_shape_type is BodyType.STANDARD:
        front_body_shape = generate_shape_of_sweater_body(data)

    def __call__(self):
        import os
        if os.path.exists(self.file_path):
            os.remove(self.file_path)
            print(f"一時ファイル {self.file_path} を削除しました。")
        else:
            print(f"一時ファイル {self.file_path} は既に存在しませんでした。")

def generate_file(data: SweaterDimensions) -> str:
    """
    データから生成した一時ファイルのパスを返す

    Args:
        data (SweaterDimensions): 検証済みの寸法データクラス

    Returns:
        str: 一時ファイルのパス
    
    """
    # TODO メイン処理
    front_body_shape = Shape.front_body_from(data)
    back_body_shape = Shape.back_body_from(data)
    sleeve_shape = Shape.sleeve_from(data)

    front_body_chart = Chart.from_shape(front_body_shape)
    back_body_chart = Chart.from_shape(back_body_shape)
    sleeve_chart = Chart.from_shape(sleeve_shape)

    start_row_of_ribbed_hem = int((data.length_of_body - data.length_of_ribbed_hem) / data.gauge.stitch_length)
    start_row_of_ribbed_cuff = int((data.length_of_sleeve - data.length_of_ribbed_cuff) / data.gauge.stitch_length)

    (
        front_body_chart
            .replace_vertical_stripes_below(front_body_chart, start_row_of_ribbed_hem, Symbol.KNIT.number, Symbol.PURL.number)
    )

    (
        back_body_chart
            .replace_vertical_stripes_below(back_body_chart, start_row_of_ribbed_hem, Symbol.KNIT.number, Symbol.PURL.number)
    )

    (
        sleeve_chart
            .replace_vertical_stripes_below(sleeve_chart, start_row_of_ribbed_cuff, Symbol.KNIT.number, Symbol.PURL.number)
    )




    tmp_file = tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.csv', encoding='utf-8', newline='')
    return tmp_file.name
    